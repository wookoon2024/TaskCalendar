from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

from taskcalendar.models import AlertType, CalendarEntry, EntryType, RecurrenceType


_HEADERS = [
    "entry_type",
    "title",
    "description",
    "day",
    "start_date",
    "end_date",
    "start_time",
    "end_time",
    "all_day",
    "assignee",
    "status",
    "recurrence_enabled",
    "recurrence_type",
    "recurrence_interval",
    "recurrence_weekdays_json",
    "recurrence_month_day",
    "recurrence_month_week",
    "recurrence_month_end",
    "completed_dates_json",
    "icon_type",
    "alert_type",
    "alert_offset",
]


def _require_openpyxl():
    try:
        from openpyxl import Workbook, load_workbook
    except Exception as exc:  # pragma: no cover
        raise RuntimeError(
            "엑셀 기능을 사용하려면 openpyxl이 필요합니다. "
            "설치 명령: python -m pip install openpyxl"
        ) from exc
    return Workbook, load_workbook


def _iso_date(value: date | None) -> str:
    return value.isoformat() if value else ""


def _parse_date_cell(value: object) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    return date.fromisoformat(text)


def _parse_bool(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "on"}


def _parse_int(value: object, fallback: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return fallback


def _parse_json_list(value: object) -> list:
    if value in (None, ""):
        return []
    if isinstance(value, list):
        return value
    parsed = json.loads(str(value))
    return parsed if isinstance(parsed, list) else []


def _strip_html_to_plain_text(text: str) -> str:
    if not text:
        return ""
    if not text.strip().startswith("<"):
        return text
    import re
    # Remove script and style elements
    clean = re.sub(r'<(script|style)\b[^>]*>([\s\S]*?)<\/\1>', '', text, flags=re.IGNORECASE)
    # Replace common block elements with newlines
    clean = re.sub(r'</?(p|div|tr|h[1-6])\b[^>]*>', '\n', clean, flags=re.IGNORECASE)
    clean = re.sub(r'<br\s*/?>', '\n', clean, flags=re.IGNORECASE)
    # Strip remaining HTML tags
    clean = re.sub(r'<[^>]+>', '', clean)
    # Decode basic entities
    clean = clean.replace('&nbsp;', ' ').replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&').replace('&quot;', '"')
    return "\n".join(line.strip() for line in clean.splitlines() if line.strip())


def export_entries_to_excel(file_path: Path, entries: list[CalendarEntry]) -> int:
    Workbook, _ = _require_openpyxl()
    wb = Workbook()
    ws = wb.active
    ws.title = "entries"
    ws.append(_HEADERS)

    for entry in entries:
        ws.append(
            [
                entry.entry_type.value,
                entry.title,
                _strip_html_to_plain_text(entry.description),
                _iso_date(entry.day),
                _iso_date(entry.start_date),
                _iso_date(entry.end_date),
                entry.start_time,
                entry.end_time,
                int(entry.all_day),
                entry.assignee,
                entry.status,
                json.dumps(entry.attachments, ensure_ascii=False),
                int(entry.recurrence_enabled),
                entry.recurrence_type.value,
                max(1, int(entry.recurrence_interval or 1)),
                json.dumps(entry.recurrence_weekdays, ensure_ascii=False),
                max(1, int(entry.recurrence_month_day or 1)),
                int(entry.recurrence_month_week or 1),
                int(entry.recurrence_month_end),
                json.dumps(entry.completed_dates, ensure_ascii=False),
                entry.icon_type,
                entry.alert_type.value,
                entry.alert_offset,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, name in enumerate(_HEADERS, start=1):
        width = min(40, max(10, len(name) + 2))
        ws.column_dimensions[chr(64 + idx)].width = width
    wb.save(file_path)
    return len(entries)


def import_entries_from_excel(file_path: Path) -> list[CalendarEntry]:
    _, load_workbook = _require_openpyxl()
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    first = next(rows, None)
    if not first:
        return []
    header = [str(item or "").strip() for item in first]
    mapping = {name: idx for idx, name in enumerate(header)}
    required_headers = [name for name in _HEADERS if name != "attachments_json"]
    missing = [name for name in required_headers if name not in mapping]
    if missing:
        raise ValueError(f"엑셀 형식이 올바르지 않습니다. 누락 컬럼: {', '.join(missing)}")

    result: list[CalendarEntry] = []
    for row_no, row in enumerate(rows, start=2):
        if row is None:
            continue

        def get(name: str):
            idx = mapping[name]
            return row[idx] if idx < len(row) else None

        raw_title = str(get("title") or "").strip()
        raw_desc = str(get("description") or "").strip()
        if not raw_title and not raw_desc:
            continue
        title = raw_title or raw_desc.splitlines()[0][:40] or "일정"

        entry_type_raw = str(get("entry_type") or EntryType.SCHEDULE.value).strip().lower()
        if entry_type_raw not in {EntryType.SCHEDULE.value, EntryType.TASK.value, EntryType.MEMO.value}:
            entry_type_raw = EntryType.SCHEDULE.value
        recurrence_raw = str(get("recurrence_type") or RecurrenceType.NONE.value).strip().lower()
        if recurrence_raw not in {item.value for item in RecurrenceType}:
            recurrence_raw = RecurrenceType.NONE.value
        alert_raw = str(get("alert_type") or AlertType.NONE.value).strip().lower()
        if alert_raw not in {item.value for item in AlertType}:
            alert_raw = AlertType.NONE.value

        try:
            entry = CalendarEntry(
                entry_type=EntryType(entry_type_raw),
                title=title,
                description=raw_desc,
                day=_parse_date_cell(get("day")),
                start_date=_parse_date_cell(get("start_date")),
                end_date=_parse_date_cell(get("end_date")),
                start_time=str(get("start_time") or "").strip(),
                end_time=str(get("end_time") or "").strip(),
                all_day=_parse_bool(get("all_day")),
                assignee=str(get("assignee") or "").strip(),
                status=str(get("status") or "").strip(),
                # Cross-device share via Excel intentionally excludes attachment paths.
                attachments=[],
                recurrence_enabled=_parse_bool(get("recurrence_enabled")),
                recurrence_type=RecurrenceType(recurrence_raw),
                recurrence_interval=max(1, _parse_int(get("recurrence_interval"), 1)),
                recurrence_weekdays=[int(v) for v in _parse_json_list(get("recurrence_weekdays_json")) if str(v).isdigit()],
                recurrence_month_day=max(1, _parse_int(get("recurrence_month_day"), 1)),
                recurrence_month_week=_parse_int(get("recurrence_month_week"), 1),
                recurrence_month_end=_parse_bool(get("recurrence_month_end")),
                completed_dates=[str(item) for item in _parse_json_list(get("completed_dates_json"))],
                icon_type=str(get("icon_type") or "").strip(),
                alert_type=AlertType(alert_raw),
                alert_offset=str(get("alert_offset") or "at_start").strip() or "at_start",
            )
        except Exception as exc:
            raise ValueError(f"{row_no}행 처리 중 오류가 발생했습니다: {exc}") from exc

        if entry.entry_type != EntryType.MEMO and entry.day is None:
            entry.day = entry.start_date or date.today()
        result.append(entry)
    return result
